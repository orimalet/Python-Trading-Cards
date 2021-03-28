# Name: Orion Assefaw, COMP517 Assignment 5(CA5)

import openpyxl

class Card():
    """
    Card class allows creating Card objects based on the following attributes: Name, Type, Health Points, 
    Moves(within which embedded is the damage factors), and Shiny Status.
    """    
    def __init__(self, theName, theType, theHP, theMoves, isShiny):
        """
        Card Class constructor function.
        Parameters:
            theName: string - The Card Name
            theType: string - The Card Type (one of the following 7 : "Magi", "Water", "Fire", "Earth", "Air", "Astral")
            theHP: int - The Card's Maximum Health Points
            theMoves: list - A list of the card's move names with their respective damage factors
            isShiny: int - The Card's shiny status (represented by 1 - for Shiny Cards and 0 - for Non-Shiny Cards)          
        Returns:
            Nothing
        """
        self.name = theName
        self.type = theType
        self.healthPoints = theHP
        self.move = theMoves
        self.shiny = isShiny
        self.shinyStatus = ""
       
    def __eq__(self, other):
        """
        A customized method for comparing two objects of a Card Class. Based on this method, two card objects are only
        deemed to be equal if they have same addresses and attributes.
        """
        return (self is other) and (self.name == other.name) and (self.type == other.type) and (self.healthPoints == other.healthPoints) and (self.move == other.move) and (self.shiny == other.shiny)
    
    def __str__(self):
        """
        A function that returns a string representation of each Card object. It includes Card's Name,Type, and Max Health Points.

        Parameters:
            Nothing

        Returns:
            Nothing
        """        
        if (self.shiny == 0):
            self.shinyStatus = "Non-Shiny"
        else:
            self.shinyStatus = "Shiny"       
        return "A {self.shinyStatus} Card {self.name}, of type {self.type}, with {self.healthPoints} health points.".format(self=self)
   
   
class Deck():
    """
    Deck class allows the general management of the Card objects such us adding cards, removing cards,viewing card's
    information, saving card's details to excel file and etc. It utilises different data structures such as lists, ints, 
    and dictionaries to carry information on the number of shiny cards in the deck, the averages of the cards in the deck, etc.
    """    
    __serialNo = 1
    def __init__(self):
        """
        Deck Class constructor function.
        Parameters:
            nothing          
        Returns:
            Nothing
        """
        self.cards = []
        self.shinyNum = 0 
        
        self.shinyCards = []
        self.overallAvg  = 0 
        self.avgsDict = {}

    def __str__(self):
        """
        A function that returns a string representation of the Deck objects. It includes total number of Cards in the Deck,
        number of Shiny Cards in the Deck and the value of the average damage factor over the enire deck.

        Parameters:
            Nothing

        Returns:
            Nothing
        """           
        self.overallAvg = (self.getAverageDamage())
        return 'The total number of cards in the Deck is {self.totalNum}. Of these, there are {self.shinyNum} Shiny Cards.\nAnd the average Damage value over the entire deck is: {self.overallAvg}\n'.format(self=self)
    
    def inputFromFile(self, fileName):
        """
        A function that inputs and populates the Deck with the cards that are present in an external excel file.

        Parameters:
            fileName: string - The name of the .xlsx excel file where cards information is to be retrieved from

        Returns:
            Nothing
        """   
        try:
            book = openpyxl.load_workbook(fileName)
            sheet = book.active
        
            nameList = []
            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):
                for cell in row:
                    if(isinstance(cell.value, str)):                
                        nameList.append(cell.value)  
       
            typeList = []
            for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
                for cell in row:
                    typeList.append(cell.value)

            hpList = []
            for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
                for cell in row:
                    hpList.append(cell.value)

            movesList = []
            for row in sheet.iter_rows(min_row=2, min_col=5, max_col=14):
                rowList = []
                for cell in row:
                    if cell.value is not None:
                        rowList.append(cell.value)
                            
                movesList.append(rowList)  
            
            shinyList = []
            for row in sheet.iter_rows(min_row=2, min_col=4, max_col=4):
                for cell in row:
                    shinyList.append(cell.value) 
                
            for i in range(len(nameList)):
                n = nameList[i]
                t = typeList[i]
                hp = hpList [i]
                m = movesList[i]
                s = shinyList[i]                  
                self.cards.append(Card(n,t,hp,m,s))
                self.totalNum = Deck.__serialNo 
                Deck.__serialNo += 1
                if s == 1:
                    self.shinyNum += 1
                
                avg = self.avgFinder(m)
                self.avgsDict[id(self.cards[i])] = avg
        
        except (TypeError , IndexError, ValueError) as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)
        
        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)
      
    def avgFinder(self,theList):
        """
        A helper function that computes an average damage across the moves of the cards in the Deck.

        Parameters:
            theList: The moves list which contains moves names and damage factors

        Returns:
            average: Average of the damages inflicted by the card moves
        """
        try:  
            damageFactors = []
            sum = 0
            for i in range(1,len(theList),2):
                damageFactors.append(theList[i])
                sum += theList[i]
            avg = sum / len(damageFactors)
            return (avg)

        except (IndexError,TypeError,NameError) as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)

        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)   
        
    def addCard(self, theCard):
        """
        A function that adds a card to the Deck.

        Parameters:
            theCard: A card that will be added to the deck

        Returns:
            Nothing
        """ 
        try:
            self.cards.append(theCard)
            self.totalNum += 1
            if(theCard.shiny == 1):
                self.shinyNum += 1
            
            avg = self.avgFinder(theCard.move)
            self.avgsDict[id(theCard)] = avg       
            self.overallAvg = self.getAverageDamage()

        except (IndexError,TypeError,NameError) as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)

        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)   

    def rmCard(self, theCard):
        """
        A function that removes a card from the Deck, if the card already exists in the deck.

        Parameters:
            theCard: A card that will be removed from the deck

        Returns:
            Nothing
        """  
        try:
            for c in self.cards:
                if((theCard == c) and (theCard is c) and (theCard.shiny == 1)):
                    self.cards.remove(c)
                    self.shinyNum -= 1
                    self.totalNum -= 1
                    del self.avgsDict[id(theCard)]
                    self.overallAvg = self.getAverageDamage()
                                       
                elif((theCard == c) and (theCard is c) and (theCard.shiny == 0)):
                    self.cards.remove(c)
                    self.totalNum -= 1
                    del self.avgsDict[id(theCard)]
                    self.overallAvg = self.getAverageDamage()

                else:
                    pass
        except ValueError as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)
            
        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)   
           
    def getMostPowerful(self):
        """
        A function that returns the card with the highest damages factor average in the entire deck.

        Parameters:
            nothing

        Returns:
            Card: most powerful card
        """ 
        try:
            avgsList = []
            for i in range (len(self.cards)):
                avgsList.append(self.avgsDict[id(self.getCards()[i])])
            
            maxAvg = avgsList[0]
            for i in range(len(avgsList)):
                if(avgsList[i] > maxAvg):
                    maxAvg = avgsList[i]
                else:
                    pass
            
            for i in range(len(self.cards)):
                if (self.avgsDict[id(self.getCards()[i])] == maxAvg):
                    return self.getCards()[i]

        except (TypeError, KeyError)  as e:
             if hasattr(e, 'message'):
                 print(e.message)
             else:
                print(e)
            
        except Exception as e:
             if hasattr(e, 'message'):
                 print(e.message)
             else:
                print(e)   

    def getAverageDamage(self):
        """
        A function that returns the average damage inflicted by the cards in the entire deck.

        Parameters:
            nothing

        Returns:
            deck average damage: as float value rounded to one decimal place
        """ 
        try:
            sum=0
            for val in self.avgsDict.values():
                sum += val
            deckAvg = sum/len(self.cards)
            return round(deckAvg,1)
                   
        except TypeError as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)
            
        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)        

    def viewAllCards(self):
        """
        A function that prints the information of all the cards in the deck. It displays the Card's Name, Type, Health Points,
        Number of Moves, and Average damage across the moves.

        Parameters:
            nothing

        Returns:
            nothing
        """
        try:
            print("There are a total of ", self.totalNum, "Cards in this Deck.")
            for i in range(len(self.cards)):            
                print("\033[4mCard\033[0m",i+1,":-","\033[4mName\033[0m:",self.getCards()[i].name,", \033[4mType\033[0m:",self.getCards()[i].type,", \033[4mHealth Points\033[0m:", self.getCards()[i].healthPoints,", \033[4mNumber of Moves\033[0m:", int(len(self.getCards()[i].move)/2),", \033[4mAverage Damage\033[0m:",self.avgsDict[id(self.getCards()[i])])
            print("")

        except (IndexError,TypeError,NameError) as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)

        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)   
        
    def viewAllShinyCards(self):
        """
        A function that prints the information of the shiny cards in the deck. 

        Parameters:
            nothing

        Returns:
            nothing
        """
        try:
            print("There is a total of", self.shinyNum, "Shiny Cards in this Deck.")
            for i in range(len(self.cards)):
                for j in range(len(self.cards[i].move)):
                    if(self.cards[i].shiny == 1):
                        print("\033[4mShiny Card Name\033[0m:",self.cards[i].name,", \033[4mType\033[0m:",self.cards[i].type,", \033[4mHealth Points\033[0m:", self.cards[i].healthPoints,", \033[4mNumber of Moves\033[0m:", int(len(self.cards[i].move)/2),", \033[4mAverage Damage\033[0m:",self.avgsDict[id(self.cards[i])])
                        break
            print("")

        except (IndexError,TypeError,NameError) as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)

        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)                    
            else:
                print(e)                 

    def viewAllByType(self, theType):
        """
        A function that prints the information of the cards that belong to the type specified by the parameter. 

        Parameters:
            theType: string - A specification of the Card Type

        Returns:
            nothing
        """
        try:
            typesList = ["Magi", "Water", "Fire", "Earth", "Air", "Astral"]
            if(theType not in typesList):
                print("No Trading card with type",theType,"exists in Tray Ding Trading Co. cards.\n")            
            else:
                count = 0
                for i in range(len(self.cards)):
                    if((theType == self.getCards()[i].type)):
                        count += 1            
                if (count>=1) :
                    print("The cards of type",theType,"are:")
                    theTypeList = []
                    for c in self.cards:
                        if c.type == theType:
                            theTypeList.append(c)
                    for c in theTypeList:
                        print("\033[4mCard Name\033[0m:",c.name, ", \033[4mHealth Points\033[0m:", c.healthPoints,", \033[4mNumber of Moves\033[0m:", int(len(c.move)/2),", \033[4mAverage Damage\033[0m:",self.avgsDict[id(c)])
                else:
                    print("No card with the type", theType,"present in this deck")
            
            print("")
        
        except (IndexError,TypeError,NameError) as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)

        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)                    
            else:
                print(e)                 
    
    def getCards(self):
        """
        A function that returns all the cards held within the deck as a collection.

        Parameters:
            nothing

        Returns:
            list - a collection of the cards in the deck
        """
        return self.cards
        
         
    def saveToFile(self, filename):
        """
        A function that saves the Deck of cards to an external excel workbook. 

        Parameters:
            filename: string - The name of the excel workbook to which the present deck of cards is saved in

        Returns:
            nothing
        """
        try:
            newBook = openpyxl.Workbook()
            sheet = newBook.active       
            title = ("Name", "Type", "HP", "Shiny", "Move Name 1", "Damage 1", "Move Name 2", "Damage 2", "Move Name 3", "Damage 3", "Move Name 4", "Damage 4", "Move Name 5", "Damage 5")
            sheet.append(title)  

            allCards = []
            for i in range(len(self.cards)):
                subCards = []
                
                subCards.append(self.cards[i].name)
                subCards.append(self.cards[i].type)
                subCards.append(self.cards[i].healthPoints)            
                subCards.append(self.cards[i].shiny)
                for j in range(len(self.cards[i].move)):
                    subCards.append(self.cards[i].move[j])
                allCards.append(subCards)
            
            for card in allCards:
                sheet.append(card)
            newBook.save(filename)

        except (IndexError,TypeError,NameError) as e:
            if hasattr(e, 'message'):
                print(e.message)
            else:
                print(e)

        except Exception as e:
            if hasattr(e, 'message'):
                print(e.message)                    
            else:
                print(e)          
